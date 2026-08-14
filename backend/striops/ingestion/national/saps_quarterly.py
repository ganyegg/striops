"""SAPS quarterly crime workbooks — station-level counts, straight from the source.

SAPS publishes one spreadsheet per quarter at
https://www.saps.gov.za/services/crimestats.php. The public sheets carry only
national and top-30 summaries, but each workbook also contains a **hidden
``RAW Data`` sheet** with a row per station per crime category and a column per
month. That is the only official machine-readable station-level series.

Each workbook covers its own quarter across five financial years, so a
continuous monthly series needs one workbook per quarter. This module fetches
and parses them; ``saps_crime`` owns the municipality aggregation and the
series that Striops serves.

Why this exists: the previous extract (``afrith/crime-stats``) stopped at
September 2025 when its maintainer stopped updating it, leaving Cape Town crime
eleven months stale on the Pulse. Reading SAPS directly removes that
dependency.
"""
from __future__ import annotations

import datetime
import re
from collections import defaultdict
from datetime import date
from io import BytesIO

import httpx

from striops.core.logging import get_logger

log = get_logger("striops.ingestion.national.saps_quarterly")

_ORDINAL = {1: "1st", 2: "2nd", 3: "3rd", 4: "4th"}

# The workbook is ~11MB, and SAPS rejects requests without a browser agent.
_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
    )
}

_RAW_SHEET = "RAW Data"


def _ssl_context() -> object:
    """Verify against the OS trust store when possible, else certifi.

    Corporate TLS-inspecting proxies (Netskope, Zscaler) re-sign responses with
    a root that is installed in the operating system but absent from certifi,
    so an httpx default would fail on a developer machine while working in
    production. Prefer the OS store and fall back rather than ever disabling
    verification.
    """
    try:
        import truststore

        return truststore.SSLContext(__import__("ssl").PROTOCOL_TLS_CLIENT)
    except Exception:  # pragma: no cover - depends on the host environment
        return True

# Column positions in the RAW Data sheet, which have been stable across the
# FY2025/26 releases. Monthly columns are found by their date header rather
# than by position, so only these anchors need to hold.
_COL_COMP_LEVEL = 2
_COL_STATION = 4
_COL_PROVINCE = 6
_COL_CODE = 8


def quarter_url(fy_start: int, quarter: int) -> str:
    """URL of the SAPS workbook for a quarter of the financial year starting ``fy_start``.

    SAPS financial years run 1 April to 31 March, so quarter 4 of FY2025/26
    covers January to March 2026.
    """
    if quarter not in _ORDINAL:
        raise ValueError(f"quarter must be 1-4, got {quarter}")
    return (
        f"https://www.saps.gov.za/services/downloads/{fy_start}/"
        f"{fy_start}-{fy_start + 1}_-_{_ORDINAL[quarter]}_Quarter_WEB.xlsx"
    )


def normalise_station(name: object) -> str:
    """Fold a station name to a join key.

    SAPS and the station register disagree on punctuation — "Gordon's Bay"
    against "Gordons Bay" — so apostrophes are deleted rather than turned into
    a space, and every other separator collapses to one.
    """
    text = str(name or "").lower().strip().replace("&", "and")
    text = re.sub(r"['\u2018\u2019`]", "", text)
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", text)).strip()


def fetch_workbook(fy_start: int, quarter: int, timeout: float = 300.0) -> bytes:
    url = quarter_url(fy_start, quarter)
    with httpx.Client(
        timeout=timeout, follow_redirects=True, headers=_HEADERS, verify=_ssl_context()
    ) as client:
        resp = client.get(url)
        resp.raise_for_status()
    log.info(
        "saps workbook fetched",
        extra={"context": {"fy": fy_start, "quarter": quarter, "bytes": len(resp.content)}},
    )
    return resp.content


def parse_station_counts(
    workbook: bytes,
    station_keys: set[str],
) -> dict[str, dict[date, float]]:
    """Aggregate monthly counts for the given stations, keyed by SAPS crime code.

    Returns ``{crime_code: {period: count}}``. Only rows at station comparison
    level are summed, so district, provincial and national roll-ups in the same
    sheet cannot double-count.
    """
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(workbook), read_only=True, data_only=True)
    if _RAW_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"{_RAW_SHEET!r} sheet not in workbook (sheets: {wb.sheetnames}) — "
            "SAPS may have changed the release format"
        )
    ws = wb[_RAW_SHEET]
    rows = ws.iter_rows(values_only=True)

    # Two banner rows precede the header row carrying the month dates.
    header: tuple | None = None
    for _ in range(6):
        candidate = next(rows, None)
        if candidate is None:
            break
        if any(isinstance(cell, datetime.datetime) for cell in candidate):
            header = candidate
            break
    if header is None:
        raise RuntimeError("no month columns found in RAW Data header")

    month_cols = {
        i: cell.date().replace(day=1)
        for i, cell in enumerate(header)
        if isinstance(cell, datetime.datetime)
    }

    totals: dict[str, dict[date, float]] = defaultdict(lambda: defaultdict(float))
    matched: set[str] = set()
    for row in rows:
        if not row or row[_COL_COMP_LEVEL] != "Station":
            continue
        key = normalise_station(row[_COL_STATION])
        if key not in station_keys:
            continue
        matched.add(key)
        code = str(row[_COL_CODE])
        for i, period in month_cols.items():
            value = row[i] if i < len(row) else None
            if isinstance(value, int | float):
                totals[code][period] += float(value)

    missing = station_keys - matched
    if missing:
        log.warning(
            "saps stations unmatched",
            extra={"context": {"count": len(missing), "stations": sorted(missing)[:10]}},
        )
    log.info(
        "saps workbook parsed",
        extra={
            "context": {
                "stations_matched": len(matched),
                "codes": len(totals),
                "months": [str(p) for p in sorted(month_cols.values())],
            }
        },
    )
    return {code: dict(periods) for code, periods in totals.items()}


def province_station_names(workbook: bytes, province: str) -> set[str]:
    """Normalised station names in one province — used to audit the join."""
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(workbook), read_only=True, data_only=True)
    ws = wb[_RAW_SHEET]
    return {
        normalise_station(row[_COL_STATION])
        for row in ws.iter_rows(min_row=4, values_only=True)
        if row and row[_COL_COMP_LEVEL] == "Station" and row[_COL_PROVINCE] == province
    }
